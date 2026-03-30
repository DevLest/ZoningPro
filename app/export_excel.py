"""Export LC application row to xlsx (LC_Application + SUM-style)."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from app.doc_requirements import format_doc_requirements_for_export
from app.models import LCApplication


def export_lc_workbook(rows: list[LCApplication], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "LC_Application"
    headers = [
        "LC-Ctrl-No.",
        "Date of Application",
        "Name of Applicant",
        "Address",
        "Name of Project",
        "Project location",
        "Doc. Requirements",
        "Lc Status",
        "Date Granted",
        "LC Fees",
        "ZC Fees",
        "Surcharge",
        "Total",
        "Category",
        "Template",
        "Project cost",
        "Lot area (sqm)",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)

    for r, app in enumerate(rows, 2):
        ws.cell(row=r, column=1, value=app.lc_ctrl_no)
        ws.cell(row=r, column=2, value=app.date_of_application)
        ws.cell(row=r, column=3, value=app.applicant_display_name)
        ws.cell(row=r, column=4, value=app.address)
        ws.cell(row=r, column=5, value=app.project_name)
        ws.cell(row=r, column=6, value=app.project_location)
        ws.cell(row=r, column=7, value=format_doc_requirements_for_export(app.doc_requirements))
        ws.cell(row=r, column=8, value=app.lc_status)
        ws.cell(row=r, column=9, value=app.date_granted)
        ws.cell(row=r, column=10, value=app.lc_fees)
        ws.cell(row=r, column=11, value=app.zc_fees)
        ws.cell(row=r, column=12, value=app.surcharge)
        ws.cell(row=r, column=13, value=app.total)
        ws.cell(row=r, column=14, value=app.category)
        ws.cell(row=r, column=15, value=app.template_id)
        ws.cell(row=r, column=16, value=app.project_cost)
        ws.cell(row=r, column=17, value=app.lot_area_sqm)

    ws2 = wb.create_sheet("SUM_export")
    sum_headers = ["CTRL #", "CLASS", "APPLICANT", "ADDRESS", "LOCATION", "AMOUNT"]
    for col, h in enumerate(sum_headers, 1):
        ws2.cell(row=1, column=col, value=h).font = Font(bold=True)

    class_map = {
        "residential": "R",
        "apartment": "A",
        "dormitory": "D",
        "commercial": "C",
        "institutional": "I",
        "special_use": "S",
    }
    for r, app in enumerate(rows, 2):
        cls = class_map.get(app.category or "", "")
        ws2.cell(row=r, column=1, value=app.lc_ctrl_no)
        ws2.cell(row=r, column=2, value=cls)
        ws2.cell(row=r, column=3, value=app.applicant_display_name)
        ws2.cell(row=r, column=4, value=app.address)
        ws2.cell(row=r, column=5, value=app.project_location or "")
        ws2.cell(row=r, column=6, value=app.total)

    for column in ws.columns:
        for cell in column:
            if cell.value is not None:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(path)
